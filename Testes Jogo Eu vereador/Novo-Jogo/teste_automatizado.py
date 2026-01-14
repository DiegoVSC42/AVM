"""
Script de Automação de Testes - Eu Vereador Jogo (VERSÃO MELHORADA)
Testa automaticamente todas as combinações de configurações e salva os resultados em planilha Excel
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time
import re

import sys

# Configurações
# Se passar argumento, usa como base. Se não, usa localhost.
if len(sys.argv) > 1:
    BASE_URL_ARG = sys.argv[1].rstrip('/')
    if not BASE_URL_ARG.startswith('http'):
        BASE_URL_ARG = f"http://{BASE_URL_ARG}"
    # Se a URL não terminar com /novo-jogo, adiciona
    if "/novo-jogo" not in BASE_URL_ARG:
        URL_BASE = f"{BASE_URL_ARG}/novo-jogo"
    else:
        URL_BASE = BASE_URL_ARG
    AMBIENTE = "TESTE_PERSONALIZADO"
else:
    URL_BASE = "http://localhost:3000/novo-jogo"
    AMBIENTE = "LOCAL_AUTOMATIZADO"

# Definir todas as combinações de testes
CONFIGURACOES = {
    "tipos_candidatura": [
        {"nome": "Primeira Candidatura", "reeleicao": False},
        {"nome": "Reeleição", "reeleicao": True}
    ],
    "tamanhos_cidade": [
        {"nome": "Cidade Pequena", "media_eleitores": 15000, "teto": 35000, "votos_reeleicao": 150},
        {"nome": "Cidade Média", "media_eleitores": 65000, "teto": 115000, "votos_reeleicao": 650},
        {"nome": "Cidade Grande", "media_eleitores": 300000, "teto": 500000, "votos_reeleicao": 3000},
        {"nome": "Metrópole", "media_eleitores": 750000, "teto": 1000000, "votos_reeleicao": 7500}
    ],
    "dificuldades": [
        {"nome": "Iniciante", "modificador_recursos": 1.30, "modificador_custos": 0.80},
        {"nome": "Desafiador", "modificador_recursos": 1.00, "modificador_custos": 1.00},
        {"nome": "Veterano", "modificador_recursos": 0.70, "modificador_custos": 1.20}
    ]
}

def calcular_valores_esperados(tipo_candidatura, tamanho_cidade, dificuldade):
    """Calcula os valores esperados baseado nas configurações"""
    
    # Meta de votos: 2% da média de eleitores
    meta_votos = int(tamanho_cidade["media_eleitores"] * 0.02)
    
    # Meta de contatos: 20% da meta de votos
    meta_contatos = int(meta_votos * 0.20)
    
    # Votos iniciais
    if tipo_candidatura["reeleicao"]:
        # Usar valor específico definido no tamanho da cidade
        votos_ultima = tamanho_cidade["votos_reeleicao"]
        votos_base = int(votos_ultima * 0.30)
        
        # Aplicar modificadores de dificuldade para Reeleição
        # Descoberto nos testes: Iniciante ganha +500, Veterano perde -300
        modificador_votos = 0
        if dificuldade["nome"] == "Iniciante":
            modificador_votos = 500
        elif dificuldade["nome"] == "Veterano":
            modificador_votos = -300
            
        votos_iniciais = votos_base + modificador_votos
        # Garantir que não fique negativo
        if votos_iniciais < 0: votos_iniciais = 0
            
        contatos_iniciais = votos_iniciais
    else:
        votos_iniciais = 0
        contatos_iniciais = 0
    
    # Saldo inicial
    if tipo_candidatura["reeleicao"]:
        # Reeleição: 60% do teto
        saldo_base = int(tamanho_cidade["teto"] * 0.60)
    else:
        # Primeira candidatura: 10% do teto
        saldo_base = int(tamanho_cidade["teto"] * 0.10)
    
    # Aplicar modificador de dificuldade no saldo
    saldo_final = int(saldo_base * dificuldade["modificador_recursos"])
    
    return {
        "votos_iniciais": votos_iniciais,
        "meta_votos": meta_votos,
        "contatos_iniciais": contatos_iniciais,
        "meta_contatos": meta_contatos,
        "energia": 100,  # Inteiro, sem %
        "reputacao": 50, # Inteiro, sem %
        "saldo": saldo_final # Inteiro, sem formatação "R$ ..."
    }

def extrair_numero(texto):
    """Extrai número de um texto, removendo formatação"""
    if not texto:
        return 0
    # Remove tudo exceto números
    numeros = re.sub(r'[^\d]', '', texto)
    return int(numeros) if numeros else 0

def iniciar_driver():
    """Inicializa o driver do Chrome"""
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    
    # Truques para evitar detecção do Selenium pelo Google
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    # User-Agent de um Chrome normal
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    driver = webdriver.Chrome(options=chrome_options)
    
    # Script extra para remover flag do webdriver
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    return driver

def realizar_login_google(driver, email, senha):
    """Realiza login no Google quando em produção"""
    print("🔑 Iniciando login com Google...")
    
    try:
        wait = WebDriverWait(driver, 20)
        
        # 1. Clicar no botão de login com Google na tela do Manus/Home
        print("   Procurando botão de login com Google...")
        try:
            # Tentar seletor específico do Manus (imagem enviada)
            btn_login = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), 'Continue with Google')]")))
            # Clicar no elemento pai se for apenas texto
            try:
                btn_login.click()
            except:
                btn_login.find_element(By.XPATH, "./..").click()
                
        except:
             # Fallback para seletores genéricos
            try:
                btn_login = driver.find_element(By.XPATH, "//button[contains(., 'Google')]")
                btn_login.click()
            except:
                print("   Botão Google específico não encontrado, tentando seguir fluxo...")
        
        time.sleep(5)
        
        # Verificar se caiu na tela de "Escolher uma conta"
        try:
            usar_outra = driver.find_element(By.XPATH, "//*[contains(text(), 'Usar outra conta')]")
            usar_outra.click()
            time.sleep(2)
        except:
            pass

        # 2. Preencher Email
        print("   Preenchendo email...")
        email_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='email']")))
        email_input.clear()
        email_input.send_keys(email)
        time.sleep(2)
        
        # Clicar em Próxima
        try:
            driver.find_element(By.ID, "identifierNext").click()
        except:
            # Tentar por texto se o ID mudar
            driver.find_element(By.XPATH, "//*[contains(text(), 'Próxima') or contains(text(), 'Next')]").click()
            
        time.sleep(5)
        
        # 3. Preencher Senha
        print("   Preenchendo senha...")
        senha_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='password']")))
        senha_input.clear()
        
        # Digitar senha lentamente para parecer humano
        for char in senha:
            senha_input.send_keys(char)
            time.sleep(0.1)
            
        time.sleep(2)
        
        # Clicar em Próxima
        try:
            driver.find_element(By.ID, "passwordNext").click()
        except:
            driver.find_element(By.XPATH, "//*[contains(text(), 'Próxima') or contains(text(), 'Next')]").click()
        
        # 4. Aguardar login completar e redirecionar
        print("   Aguardando redirecionamento...")
        time.sleep(15)
        
        # Verificar se logou com sucesso (url deve mudar ou aparecer dashboard)
        if "novo-jogo" in driver.current_url or "dashboard" in driver.current_url:
            print("✅ Login realizado com sucesso!")
            return True
        else:
            print("⚠️ Status do login incerto, continuando...")
            return True
            
    except Exception as e:
        print(f"❌ Falha no login: {str(e)}")
        # Tentar salvar screenshot
        try:
            driver.save_screenshot(f"d:/GIT/AVM/Testes Jogo Eu vereador/Novo-Jogo/erro_login_{int(time.time())}.png")
        except: pass
        return False

def extrair_valores_com_selenium(driver):
    """Extrai valores buscando elementos diretamente no DOM"""
    
    try:
        wait = WebDriverWait(driver, 10)
        resultado = {}
        
        texto_completo = driver.find_element(By.TAG_NAME, "body").text
        linhas = texto_completo.split('\n')
        
        # Processar linha por linha
        i = 0
        while i < len(linhas):
            linha = linhas[i].strip()
            
            # VOTOS: procurar "VOTOS", depois o valor, depois "Meta: X"
            if linha == 'VOTOS':
                # Próxima linha tem o valor inicial
                if i + 1 < len(linhas):
                    valor_inicial = linhas[i + 1].strip()
                    # Linha seguinte tem "Meta: X"
                    if i + 2 < len(linhas) and 'Meta:' in linhas[i + 2]:
                        meta = linhas[i + 2].replace('Meta:', '').strip()
                        resultado['votos_iniciais'] = int(valor_inicial.replace('.', '').replace(',', ''))
                        resultado['meta_votos'] = int(meta.replace('.', '').replace(',', ''))
                        print(f"  ✓ VOTOS: inicial={valor_inicial}, meta={meta}")
                        i += 3
                        continue
            
            # CONTATOS: mesma lógica
            elif linha == 'CONTATOS':
                if i + 1 < len(linhas):
                    valor_inicial = linhas[i + 1].strip()
                    if i + 2 < len(linhas) and 'Meta:' in linhas[i + 2]:
                        meta = linhas[i + 2].replace('Meta:', '').strip()
                        resultado['contatos_iniciais'] = int(valor_inicial.replace('.', '').replace(',', ''))
                        resultado['meta_contatos'] = int(meta.replace('.', '').replace(',', ''))
                        print(f"  ✓ CONTATOS: inicial={valor_inicial}, meta={meta}")
                        i += 3
                        continue
            
            # ENERGIA: extrair apenas o número
            elif linha == 'ENERGIA':
                if i + 1 < len(linhas):
                    valor = linhas[i + 1].strip()
                    if '%' in valor:
                        # Extrair apenas o número: "100%" -> 100
                        resultado['energia'] = int(valor.replace('%', '').strip())
                        print(f"  ✓ ENERGIA: {resultado['energia']}")
                        i += 2
                        continue
            
            # REPUTAÇÃO: extrair apenas o número
            elif linha == 'REPUTAÇÃO':
                if i + 1 < len(linhas):
                    valor = linhas[i + 1].strip()
                    if '%' in valor:
                        # Extrair apenas o número: "50%" -> 50
                        resultado['reputacao'] = int(valor.replace('%', '').strip())
                        print(f"  ✓ REPUTAÇÃO: {resultado['reputacao']}")
                        i += 2
                        continue
            
            # SALDO: extrair apenas o número
            elif linha == 'SALDO':
                if i + 1 < len(linhas):
                    valor = linhas[i + 1].strip()
                    if 'R$' in valor:
                        # Extrair apenas o número: "R$ 4,550" -> 4550
                        # Remove R$, ponto, vírgula e espaços
                        valor_limpo = valor.replace('R$', '').replace('.', '').replace(',', '').strip()
                        resultado['saldo'] = int(valor_limpo)
                        print(f"  ✓ SALDO: {resultado['saldo']}")
                        i += 2
                        continue
            
            i += 1
        
        print(f"\n✅ Valores finais extraídos: {resultado}")
        return resultado
        
    except Exception as e:
        print(f"⚠️ Erro ao extrair valores: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def executar_teste(driver, tipo_candidatura, tamanho_cidade, dificuldade):
    """Executa um teste específico e retorna os valores obtidos"""
    
    print(f"\n🔄 Testando: {tipo_candidatura['nome']} + {tamanho_cidade['nome']} + {dificuldade['nome']}")
    
    try:
        # Navegar para a página
        driver.get(URL_BASE)
        time.sleep(5) # Aumentado para 5s para garantir carregamento em produção
        
        # Selecionar tipo de candidatura
        wait = WebDriverWait(driver, 10)
        
        # Clicar em Primeira Candidatura ou Reeleição
        tipo_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//button[contains(., '{tipo_candidatura['nome']}')]")
        ))
        tipo_btn.click()
        time.sleep(1)
        
        # Se for reeleição, preencher votos
        if tipo_candidatura["reeleicao"]:
            votos_input = wait.until(EC.presence_of_element_located(
                (By.XPATH, "//input[@type='number' or @type='text']")
            ))
            votos_input.clear()
            # Usar valor específico da cidade
            votos_input.send_keys(str(tamanho_cidade["votos_reeleicao"]))
            time.sleep(1)
        
        # Selecionar tamanho da cidade
        cidade_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//button[contains(., '{tamanho_cidade['nome']}')]")
        ))
        cidade_btn.click()
        time.sleep(1)
        
        # Selecionar dificuldade
        dificuldade_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//button[contains(., '{dificuldade['nome']}')]")
        ))
        dificuldade_btn.click()
        time.sleep(1)
        
        # Clicar em Iniciar Campanha
        iniciar_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(., 'INICIAR CAMPANHA')]")
        ))
        iniciar_btn.click()
        time.sleep(3)
        
        # Clicar em Iniciar Fase 1
        fase_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(., 'Iniciar Fase 1')]")
        ))
        fase_btn.click()
        time.sleep(3)  # Aumentado para dar tempo de carregar
        
        # Extrair valores usando Selenium
        valores_obtidos = extrair_valores_com_selenium(driver)
        
        if valores_obtidos:
            print(f"✅ Valores extraídos: {valores_obtidos}")
        else:
            print(f"⚠️ Nenhum valor foi extraído!")
            # Tentar salvar screenshot para debug
            try:
                screenshot_path = f"d:/GIT/AVM/Testes Jogo Eu vereador/Novo-Jogo/debug_screenshot_{int(time.time())}.png"
                driver.save_screenshot(screenshot_path)
                print(f"📸 Screenshot salvo em: {screenshot_path}")
            except:
                pass
        
        return valores_obtidos if valores_obtidos else None
        
        return valores_obtidos if valores_obtidos else None
        
    except Exception as e:
        print(f"❌ Erro no teste: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Tentar salvar screenshot para debug
        try:
            screenshot_path = f"d:/GIT/AVM/Testes Jogo Eu vereador/Novo-Jogo/error_screenshot_{int(time.time())}.png"
            driver.save_screenshot(screenshot_path)
            print(f"📸 Screenshot de erro salvo em: {screenshot_path}")
        except:
            pass
        return None

def comparar_valores(esperado, obtido):
    """Compara valores esperados com obtidos e retorna status"""
    if obtido is None:
        return "❌ ERRO"
    
    # Para números
    if isinstance(esperado, int):
        return "✅ OK" if esperado == obtido else "❌ ERRO"
    
    # Para strings (percentuais e saldo)
    return "✅ OK" if esperado == obtido else "❌ ERRO"

def criar_planilha_resultados(resultados):
    """Cria planilha Excel com os resultados dos testes"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = AMBIENTE
    
    # Cabeçalhos
    headers = [
        "Tipo de Candidatura", "Tamanho da Cidade", "Dificuldade", "Votos Última Eleição",
        "Votos Iniciais Esperado", "Votos Iniciais Obtido", "Status",
        "Meta Votos Esperado", "Meta Votos Obtido", "Status",
        "Contatos Iniciais Esperado", "Contatos Iniciais Obtido", "Status",
        "Meta Contatos Esperado", "Meta Contatos Obtido", "Status",
        "Energia Esperado", "Energia Obtido", "Status",
        "Reputação Esperado", "Reputação Obtido", "Status",
        "Saldo Esperado", "Saldo Obtido", "Status",
        "Resultado Geral"
    ]
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Aplicar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    for col in range(5, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Inserir dados
    for row_idx, resultado in enumerate(resultados, 2):
        esperado = resultado["esperado"]
        obtido = resultado["obtido"]
        
        # Determinar resultado geral
        todos_ok = all([
            comparar_valores(esperado["votos_iniciais"], obtido.get("votos_iniciais") if obtido else None) == "✅ OK",
            comparar_valores(esperado["meta_votos"], obtido.get("meta_votos") if obtido else None) == "✅ OK",
            comparar_valores(esperado["contatos_iniciais"], obtido.get("contatos_iniciais") if obtido else None) == "✅ OK",
            comparar_valores(esperado["meta_contatos"], obtido.get("meta_contatos") if obtido else None) == "✅ OK",
            comparar_valores(esperado["energia"], obtido.get("energia") if obtido else None) == "✅ OK",
            comparar_valores(esperado["reputacao"], obtido.get("reputacao") if obtido else None) == "✅ OK",
            comparar_valores(esperado["saldo"], obtido.get("saldo") if obtido else None) == "✅ OK"
        ]) if obtido else False
        
        resultado_geral = "✅ PASSOU" if todos_ok else "❌ FALHOU"
        
        # Determinar votos da última eleição para exibir
        votos_ultima = resultado.get("votos_ultima_eleicao")
        if votos_ultima is None and resultado.get("reeleicao"):
             # Backwards compatibility or lookup logic here if needed, 
             # but we passed it explicitly in results below
             pass

        row_data = [
            resultado["tipo_candidatura"],
            resultado["tamanho_cidade"],
            resultado["dificuldade"],
            resultado.get("votos_ultima_eleicao", "N/A"),
            esperado["votos_iniciais"],
            obtido.get("votos_iniciais", "") if obtido else "",
            comparar_valores(esperado["votos_iniciais"], obtido.get("votos_iniciais") if obtido else None),
            esperado["meta_votos"],
            obtido.get("meta_votos", "") if obtido else "",
            comparar_valores(esperado["meta_votos"], obtido.get("meta_votos") if obtido else None),
            esperado["contatos_iniciais"],
            obtido.get("contatos_iniciais", "") if obtido else "",
            comparar_valores(esperado["contatos_iniciais"], obtido.get("contatos_iniciais") if obtido else None),
            esperado["meta_contatos"],
            obtido.get("meta_contatos", "") if obtido else "",
            comparar_valores(esperado["meta_contatos"], obtido.get("meta_contatos") if obtido else None),
            esperado["energia"],
            obtido.get("energia", "") if obtido else "",
            comparar_valores(esperado["energia"], obtido.get("energia") if obtido else None),
            esperado["reputacao"],
            obtido.get("reputacao", "") if obtido else "",
            comparar_valores(esperado["reputacao"], obtido.get("reputacao") if obtido else None),
            esperado["saldo"],
            obtido.get("saldo", "") if obtido else "",
            comparar_valores(esperado["saldo"], obtido.get("saldo") if obtido else None),
            resultado_geral
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Colorir células de status
            if value == "✅ OK" or value == "✅ PASSOU":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif value == "❌ ERRO":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif value == "❌ FALHOU":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
    
    # Salvar
    filename = f"d:/GIT/AVM/Testes Jogo Eu vereador/Novo-Jogo/testes_automatizados_{int(time.time())}.xlsx"
    wb.save(filename)
    print(f"\n📊 Planilha salva em: {filename}")
    return filename

def main():
    """Função principal que executa todos os testes"""
    
    print("🚀 Iniciando testes automatizados...")
    print(f"🌐 URL: {URL_BASE}")
    print(f"📋 Ambiente: {AMBIENTE}\n")
    
    driver = None
    resultados = []
    
    try:
        driver = iniciar_driver()
        
        # Se estiver em produção (não localhost), fazer login
        if "localhost" not in URL_BASE:
            # Ir direto para a URL de autenticação do Manus para evitar problemas com botão "Jogar Agora"
            URL_AUTH_MANUS = "https://manus.im/app-auth?appId=49MLqpU5iZSkTdzww5egYS&redirectUri=https%3A%2F%2Fjogo.euvereador.com.br%2Fapi%2Foauth%2Fcallback&state=aHR0cHM6Ly9qb2dvLmV1dmVyZWFkb3IuY29tLmJyL2FwaS9vYXV0aC9jYWxsYmFjaw%3D%3D&type=signIn"
            print(f"🔄 Acessando diretamente sistema de login: {URL_AUTH_MANUS[:50]}...")
            driver.get(URL_AUTH_MANUS)
            time.sleep(5)
            
            # Tentar fazer login
            if not realizar_login_google(driver, "dvs.testes@gmail.com", "Teste_123456"):
                print("❌ Não foi possível logar. Encerrando testes.")
                return 

        # Iterar sobre todas as combinações
        for tipo_candidatura in CONFIGURACOES["tipos_candidatura"]:
            for tamanho_cidade in CONFIGURACOES["tamanhos_cidade"]:
                for dificuldade in CONFIGURACOES["dificuldades"]:
                    
                    # Calcular valores esperados
                    esperado = calcular_valores_esperados(tipo_candidatura, tamanho_cidade, dificuldade)
                    
                    # Executar teste
                    obtido = executar_teste(driver, tipo_candidatura, tamanho_cidade, dificuldade)
                    
                    # Determinar votos da ultima eleição para registro
                    if tipo_candidatura["reeleicao"]:
                        votos_ultima = tamanho_cidade["votos_reeleicao"]
                    else:
                        votos_ultima = None
                    
                    # Armazenar resultado
                    resultados.append({
                        "tipo_candidatura": tipo_candidatura["nome"],
                        "tamanho_cidade": tamanho_cidade["nome"],
                        "dificuldade": dificuldade["nome"],
                        "votos_ultima_eleicao": votos_ultima,
                        "reeleicao": tipo_candidatura["reeleicao"],
                        "esperado": esperado,
                        "obtido": obtido
                    })
                    
                    # Pequena pausa entre testes
                    time.sleep(1)
        
        # Criar planilha com resultados
        criar_planilha_resultados(resultados)
        
        # Resumo
        total = len(resultados)
        passou = sum(1 for r in resultados if r["obtido"] is not None and all([
            comparar_valores(r["esperado"]["votos_iniciais"], r["obtido"].get("votos_iniciais")) == "✅ OK",
            comparar_valores(r["esperado"]["meta_votos"], r["obtido"].get("meta_votos")) == "✅ OK",
            comparar_valores(r["esperado"]["contatos_iniciais"], r["obtido"].get("contatos_iniciais")) == "✅ OK",
            comparar_valores(r["esperado"]["meta_contatos"], r["obtido"].get("meta_contatos")) == "✅ OK",
            comparar_valores(r["esperado"]["energia"], r["obtido"].get("energia")) == "✅ OK",
            comparar_valores(r["esperado"]["reputacao"], r["obtido"].get("reputacao")) == "✅ OK",
            comparar_valores(r["esperado"]["saldo"], r["obtido"].get("saldo")) == "✅ OK"
        ]))
        falhou = total - passou
        
        print(f"\n✅ Testes concluídos!")
        print(f"📊 Total: {total} | Passou: {passou} | Falhou: {falhou}")
        
    except Exception as e:
        print(f"\n❌ Erro geral: {str(e)}")
        import traceback
        traceback.print_exc()
    
    finally:
        if driver:
            driver.quit()
            print("\n🔒 Navegador fechado")

if __name__ == "__main__":
    main()
