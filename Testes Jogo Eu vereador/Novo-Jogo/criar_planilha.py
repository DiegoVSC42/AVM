import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Criar workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove a aba padrão

# Criar abas
ws_prod = wb.create_sheet("PRODUÇÃO")
ws_local = wb.create_sheet("LOCAL")

# Cabeçalhos
headers = [
    "Tipo de Candidatura", "Tamanho da Cidade", "Dificuldade", "Votos Última Eleição",
    "Votos Iniciais Esperado", "Votos Iniciais Obtido", "Status",
    "Meta Votos Esperado", "Meta Votos Obtido", "Status",
    "Contatos Iniciais Esperado", "Contatos Iniciais Obtido", "Status",
    "Meta Contatos Esperado", "Meta Contatos Obtido", "Status",
    "Energia Esperado", "Energia Obtido", "Status",
    "Reputação Esperado", "Reputação Obtido", "Status",
    "Saldo Esperado", "Saldo Obtido", "Status",
    "Resultado Geral", "Arquivos de Erro"
]

# Estilo do cabeçalho
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Aplicar cabeçalhos nas duas abas
for ws in [ws_prod, ws_local]:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    for col in range(5, 27):
        ws.column_dimensions[get_column_letter(col)].width = 15

# Dados PRODUÇÃO
prod_data = [
    ["Primeira Candidatura", "Cidade Pequena", "Desafiador", "",
     0, 0, "✅ OK", 300, 400, "❌ ERRO", 0, 0, "✅ OK", 60, 80, "❌ ERRO",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 3.500", "R$ 10.000", "❌ ERRO", "❌ FALHOU",
     "erro_primeira_pequena_desafiador.png, .csv, .md"],
    
    ["Reeleição", "Cidade Pequena", "Desafiador", 30000,
     9000, 0, "❌ CRÍTICO", 300, 400, "❌ ERRO", 9000, 0, "❌ CRÍTICO", 60, 80, "❌ ERRO",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 21.000", "R$ 10.000", "❌ CRÍTICO", "❌ FALHOU",
     "erro_reeleicao_pequena_desafiador.png, .csv, .md"],
    
    ["Primeira Candidatura", "Cidade Média", "Desafiador", "",
     0, 0, "✅ OK", 1300, 1300, "✅ OK", 0, 0, "✅ OK", 260, 260, "✅ OK",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 11.500", "R$ 10.000", "❌ ERRO", "❌ FALHOU",
     "erro_primeira_media_desafiador.png, .csv, .md"]
]

# Dados LOCAL - Organizados por: Tamanho da Cidade -> Dificuldade (Iniciante, Desafiador, Veterano)
local_data = [
    # Cidade Pequena
    ["Primeira Candidatura", "Cidade Pequena", "Iniciante", "",
     0, 0, "✅ OK", 300, 300, "✅ OK", 0, 0, "✅ OK", 60, 60, "✅ OK",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 4.550", "R$ 4.550", "✅ OK", "✅ PASSOU", "Nenhum (teste passou)"],
    
    ["Primeira Candidatura", "Cidade Pequena", "Desafiador", "",
     0, 0, "✅ OK", 300, 300, "✅ OK", 0, 0, "✅ OK", 60, 60, "✅ OK",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 3.500", "R$ 3.500", "✅ OK", "✅ PASSOU", "Nenhum (teste passou)"],
    
    ["Primeira Candidatura", "Cidade Pequena", "Veterano", "",
     0, 0, "✅ OK", 300, 300, "✅ OK", 0, 0, "✅ OK", 60, 60, "✅ OK",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 2.450", "R$ 2.450", "✅ OK", "✅ PASSOU", "Nenhum (teste passou)"],
    
    # Cidade Média
    ["Primeira Candidatura", "Cidade Média", "Iniciante", "",
     0, 0, "✅ OK", 1300, 1300, "✅ OK", 0, 0, "✅ OK", 260, 260, "✅ OK",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 14.950", "R$ 14.950", "✅ OK", "✅ PASSOU", "Nenhum (teste passou)"],
    
    ["Primeira Candidatura", "Cidade Média", "Desafiador", "",
     0, 0, "✅ OK", 1300, 1300, "✅ OK", 0, 0, "✅ OK", 260, 260, "✅ OK",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 11.500", "R$ 11.500", "✅ OK", "✅ PASSOU", "Nenhum (teste passou)"],
    
    ["Primeira Candidatura", "Cidade Média", "Veterano", "",
     0, 0, "✅ OK", 1300, 1300, "✅ OK", 0, 0, "✅ OK", 260, 260, "✅ OK",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 8.050", "R$ 8.050", "✅ OK", "✅ PASSOU", "Nenhum (teste passou)"],
    
    # Reeleição (mantendo para referência, mas vou organizar depois)
    ["Reeleição", "Cidade Pequena", "Desafiador", 30000,
     9000, 9000, "✅ OK", 300, 300, "✅ OK", 9000, 9000, "✅ OK", 60, 60, "✅ OK",
     "100%", "100%", "✅ OK", "50%", "50%", "✅ OK",
     "R$ 21.000", "R$ 21.000", "✅ OK", "✅ PASSOU", "Nenhum (teste passou)"]
]

# Inserir dados
for row_idx, row_data in enumerate(prod_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_prod.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Colorir células de status
        if value == "✅ OK" or value == "✅ PASSOU":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif value == "❌ ERRO":
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif value == "❌ CRÍTICO" or value == "❌ FALHOU":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

for row_idx, row_data in enumerate(local_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_local.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Colorir células de status
        if value == "✅ OK" or value == "✅ PASSOU":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif value == "❌ ERRO":
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif value == "❌ CRÍTICO" or value == "❌ FALHOU":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

# Salvar
wb.save("d:/GIT/AVM/Testes Jogo Eu vereador/Novo-Jogo/consolidado_testes.xlsx")
print("Planilha criada com sucesso!")
